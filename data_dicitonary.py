###
# Genearte data dictionary based on metadata and sample value
###

from openpyxl import Workbook
from openpyxl import load_workbook


from db_connect.sqlserver_db import UseSqlserverDB
import conf.acct as acct
import tool.TSQL as TSQL_function
import tool.tool as tool
from tool.tool import pop_db_name

CURRENT_DB = acct.QA_ID_CAMPING_MART
pop_db_name(CURRENT_DB)
SEED_FILE = r'.\seed\DataDictionary_Template.xlsx'
excelName = tool.file_name('DataDictionary','xlsx')
workbook = load_workbook(SEED_FILE)

with UseSqlserverDB(CURRENT_DB) as cursor:
    sheet = workbook['DataDictionary']
    rows = sheet.rows
    columns = sheet.columns

    for row in range(2,sheet.max_row+1):
        tableName = str(sheet.cell(row=row,column=1).value)
        columnName = str(sheet.cell(row=row,column=4).value)

        #print (tableName + ":" + columnName)
        query = "SELECT TOP 1 [" + columnName + "] FROM " + tableName + " WITH(NOLOCK) WHERE [" + columnName + "] IS NOT NULL"
        result = str(TSQL_function.inquery_single_row(query,cursor))
        sheet.cell(row=row,column=12).value = result

workbook.save(excelName)  

